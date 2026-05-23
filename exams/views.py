from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.db.models import Q
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from questions.models import Exam
from .models import ExamAttempt, StudentAnswer, Grade
from .decorators import student_required
from questions.decorators import teacher_required

logger = logging.getLogger("exams")


# ================== AUTO SCORE ==================
def _compute_auto_score(exam: Exam, attempt: ExamAttempt) -> float:
    """
    Auto-score برای سوال‌های تستی:
    اگر درست بود: score همون سوال اضافه میشه
    """
    test_questions = list(
        exam.questions.filter(question_type="test").prefetch_related("choices")
    )
    if not test_questions:
        logger.info("[AUTO] No test questions found.")
        return 0.0

    # جواب‌های تستی همین attempt
    answers_map = {
        a.question_id: a.selected_choice_id
        for a in attempt.answers.select_related("question")
        if a.question.question_type == "test"
    }

    total = 0.0

    for q in test_questions:
        correct_choice = q.choices.filter(is_correct=True).first()
        selected_choice_id = answers_map.get(q.id)

        logger.info(
            f"[AUTO] Q{q.id} selected={selected_choice_id} correct={getattr(correct_choice,'id',None)} q_score={q.score}"
        )

        if not correct_choice:
            logger.warning(f"[AUTO] Q{q.id} has NO correct choice (is_correct=True).")
            continue

        if selected_choice_id is not None and int(selected_choice_id) == int(correct_choice.id):
            total += float(q.score or 0)

    return round(total, 2)


def _has_descriptive(exam: Exam) -> bool:
    return exam.questions.filter(question_type="descriptive").exists()


def _send_grade_email(student_email: str, exam_title: str, score: float, total: float):
    subject = f"Grade Result - {exam_title}"
    message = f"Your exam '{exam_title}' has been graded.\n\nScore: {score} / {total}\n"
    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", settings.EMAIL_HOST_USER)

    send_mail(subject, message, from_email, [student_email], fail_silently=True)
    logger.info(f"[EMAIL] to={student_email} subject={subject} message={message}")


# ================== STUDENT ==================
@student_required
def student_dashboard(request):
    q = request.GET.get("q", "").strip()

    exams = Exam.objects.filter(is_published=True)
    if q:
        exams = exams.filter(Q(title__icontains=q) | Q(subject__icontains=q)).distinct()

    today = timezone.localdate()
    exam_info = []

    for exam in exams:
        attempt = ExamAttempt.objects.filter(student=request.user, exam=exam).first()
        grade = Grade.objects.filter(attempt=attempt).first() if attempt else None

        # تعیین وضعیت و کلاس رنگ
        if exam.start_time.date() > today:
            status = "Not Started Yet"
            can_start = False
            status_class = "gray"
        elif exam.end_time.date() < today:
            status = "Finished"
            can_start = False
            status_class = "red"
        else:
            if attempt:
                if grade and grade.score is not None:
                    status = f"Score: {grade.score}"
                    status_class = "green"
                else:
                    status = "Submitted / Waiting for grade"
                    status_class = "orange"
                can_start = False
            else:
                status = "Start"
                can_start = True
                status_class = "blue"

        exam_info.append({
            "exam": exam,
            "status": status,
            "can_start": can_start,
            "status_class": status_class,
            "start_date": exam.start_time,
            "end_date": exam.end_time,
        })

    return render(request, "exams/student_dashboard.html", {
        "exam_info": exam_info,
        "query": q,
    })


@student_required
def start_exam(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, is_published=True)

    today = timezone.localdate()
    if not (exam.start_time.date() <= today <= exam.end_time.date()):
        return redirect("exams:student_dashboard")

    if ExamAttempt.objects.filter(student=request.user, exam=exam).exists():
        return redirect("exams:student_dashboard")

    if request.method == "POST":
        attempt = ExamAttempt.objects.create(
            student=request.user,
            exam=exam,
            is_finished=False,
        )

        for q in exam.questions.all():
            if q.question_type == "test":
                StudentAnswer.objects.create(
                    attempt=attempt,
                    question=q,
                    selected_choice_id=request.POST.get(f"question_{q.id}") or None,
                )
            else:
                StudentAnswer.objects.create(
                    attempt=attempt,
                    question=q,
                    descriptive_answer=(request.POST.get(f"text_question_{q.id}") or "").strip(),
                    uploaded_file=request.FILES.get(f"file_question_{q.id}") or None,
                )

        attempt.is_finished = True
        attempt.finished_at = timezone.now()
        attempt.save()

        if not _has_descriptive(exam):
            auto_score = _compute_auto_score(exam, attempt)

            grade, _ = Grade.objects.get_or_create(attempt=attempt)
            grade.score = auto_score
            grade.save()

            _send_grade_email(
                request.user.email,
                exam.title,
                float(grade.score or 0),
                float(exam.total_score or 0),
            )

        return redirect("exams:student_dashboard")

    return render(request, "exams/start_exam.html", {"exam": exam})


# ================== TEACHER ==================
@teacher_required
def exam_attempts(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, teacher=request.user)
    attempts = (
        ExamAttempt.objects.filter(exam=exam, is_finished=True)
        .select_related("student")
        .order_by("-started_at")
    )
    return render(request, "exams/exam_attempts.html", {"exam": exam, "attempts": attempts})


@teacher_required
def attempt_detail(request, attempt_id):
    attempt = get_object_or_404(
        ExamAttempt.objects.select_related("exam", "student"),
        id=attempt_id,
        exam__teacher=request.user,
    )
    
    # محاسبه نمره تستی (مجموع نمره سوالاتی که پاسخ صحیح دارند)
    test_score = 0.0
    for answer in attempt.answers.filter(question__question_type="test").select_related("selected_choice"):
        if answer.selected_choice and answer.selected_choice.is_correct:
            test_score += float(answer.question.score or 0)
    
    exam_total = float(attempt.exam.total_score or 0)
    
    # دریافت نمره ذخیره شده (در صورت وجود)
    grade, _ = Grade.objects.get_or_create(attempt=attempt)
    
    # نمره تشریحی ذخیره شده (اگر grade.score وجود داشته باشد، مقدار تشریحی = grade.score - test_score)
    if grade.score is not None:
        descriptive_score = max(0, float(grade.score) - test_score)
    else:
        descriptive_score = 0.0
    
    remaining = max(exam_total - test_score, 0)
    
    if request.method == "POST":
        raw = request.POST.get("score", "")
        try:
            new_desc_score = float(raw)
        except (TypeError, ValueError):
            return redirect("exams:attempt_detail", attempt_id=attempt.id)
        
        if 0 <= new_desc_score <= remaining:
            new_total = test_score + new_desc_score
            grade.score = new_total
            grade.save()
            
            # ارسال ایمیل نمره به دانشجو (اختیاری)
            _send_grade_email(
                attempt.student.email,
                attempt.exam.title,
                new_total,
                exam_total,
            )
            
            return redirect("exams:exam_attempts", exam_id=attempt.exam.id)
        else:
            # اگر نمره خارج از محدوده است، صفحه دوباره بارگذاری شود (خطا اضافه نشده)
            pass
    
    context = {
        "attempt": attempt,
        "answers": attempt.answers.select_related("question", "selected_choice"),
        "grade": grade,
        "test_score": test_score,
        "remaining_score": remaining,
        "exam_total": exam_total,
        "descriptive_score": descriptive_score,  
    }
    return render(request, "exams/attempt_detail.html", context)


@teacher_required
def grade_exam(request, attempt_id):
    return redirect("exams:attempt_detail", attempt_id=attempt_id)


# ================== EXPORT EXCEL ==================
@teacher_required
def export_attempts_to_excel(request, exam_id):
    exam = get_object_or_404(Exam, id=exam_id, teacher=request.user)
    attempts = ExamAttempt.objects.filter(exam=exam, is_finished=True).select_related("student").order_by("-started_at")
    
    # ایجاد workbook و active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{exam.title[:30]} Grades"
    
    # تعریف هدرها (سطر اول) - شامل نمره کل
    headers = ["شماره", "نام کاربری دانشجو", "نمره کسب شده", "نمره کل (حداکثر)", "تاریخ شروع", "تاریخ پایان"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # پر کردن داده‌ها
    total_score_max = float(exam.total_score or 0)
    for row_idx, attempt in enumerate(attempts, start=2):
        grade = Grade.objects.filter(attempt=attempt).first()
        obtained = grade.score if grade and grade.score is not None else "Not graded"
        ws.cell(row=row_idx, column=1, value=row_idx-1)
        ws.cell(row=row_idx, column=2, value=attempt.student.username)
        ws.cell(row=row_idx, column=3, value=obtained)
        ws.cell(row=row_idx, column=4, value=total_score_max)
        ws.cell(row=row_idx, column=5, value=attempt.started_at.strftime("%Y-%m-%d %H:%M") if attempt.started_at else "")
        ws.cell(row=row_idx, column=6, value=attempt.finished_at.strftime("%Y-%m-%d %H:%M") if attempt.finished_at else "")
    
    # تنظیم عرض ستون‌ها
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # ایجاد پاسخ HTTP با نوع فایل Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{exam.title}_grades.xlsx"'
    wb.save(response)
    return response